import openpyxl
import os
from openpyxl.styles import PatternFill, Font
import time

from fill_dictionnary_of_return_values import fill_dictionnary_of_return_values
from build_dict_functions import build_dict_functions
from global_declarations import *
from get_output_file_name import get_output_file_name

def create_excel(project_folder: str, scope_folder: str, progress_callback, output_file_path: str = None) -> None:
    """
    Main app that does everything
    :param project_folder: Folder of the project. It's the folder that contains the function that needs to be digged.
    :param scope_folder: Scope of the project, to find the reference of other functions.
    :param progress_callback: Function to emit a signal for the progress bar.
    :param output_file_path: If it's none, the output is saved in the current path.
    """

    project_folder = project_folder.replace("\\", "/")
    scope_folder = scope_folder.replace("\\", "/")

    full_dictionary = fill_dictionnary_of_return_values(project_folder, scope_folder, progress_callback)
    
    #(full_dictionary, full_list) = build_dict_functions(project_folder)

    classeur = openpyxl.Workbook()
    feuille = classeur.active

    number_of_line = 1

    couleurs_pastels = ['FFFFB6C1', 'FF98FB98', 'FFFFDAB9', 'FFADD8E6', 'FFFFA07A']

    for functions_Path, functions_Info in full_dictionary.items():
        
        last_Functions_Path = functions_Path

        current_color = couleurs_pastels[0]
        object_color = openpyxl.styles.colors.Color(rgb=current_color)
        couleurs_pastels.append(couleurs_pastels.pop(0))
        
        for function in functions_Info:
            
            # We create a new line only if the function has to be analysed.
            if "Not analysed." not in function.returned_values:
                feuille[f'A{number_of_line}'] = functions_Path
                feuille[f'A{number_of_line}'].fill = PatternFill(start_color=object_color, end_color=object_color, fill_type="solid")

                feuille[f'B{number_of_line}'] = function.function_name

                if function.returned_values == []:
                    return_values = ""
                else:
                    return_values = ', '.join(map(str, function.returned_values))
                feuille[f'C{number_of_line}'] = return_values

                #If an unknown value is found, the cell will be in red.
                if "Attention" in return_values:
                    feuille[f'C{number_of_line}'].fill = PatternFill(start_color='FFB50000', end_color='FFB50000', fill_type="solid")
                    feuille[f'C{number_of_line}'].font = Font(color='FFFFFFFF')
                
                number_of_line += 1

    feuille.column_dimensions['A'].width = 100
    feuille.column_dimensions['B'].width = 50
    feuille.column_dimensions['C'].width = 150

    # Get the output file name.
    output_file_name = get_output_file_name(output_file_path)

    # Save the file.
    classeur.save(output_file_name)
    
    #print(f"The file {output_file_name} has been created.")

if __name__ == "__main__":
    pass